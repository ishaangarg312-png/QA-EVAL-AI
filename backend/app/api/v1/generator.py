import io
import json
import re
import time
import logging
import uuid
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import httpx

from app.core.database import get_db
from app.models.organization import User, AIProviderSetting
from app.models.test_case import TestDataset
from app.models.project import Project
from sqlalchemy.orm.attributes import flag_modified
from app.api.v1.auth import get_authenticated_user
from app.core.kill_switch import SystemKillSwitchManager
from app.core.security import decrypt_secret
from app.core.doc_parser import extract_text_from_file
from app.core.ai_discovery import record_llm_usage, PROVIDER_METADATA
from app.schemas.generator_schemas import (
    ColumnConfig,
    EntityLevel,
    GenerateTestRequest,
    ExportExcelRequest,
    SaveDatasetRequest,
    SavePromptRequest,
    SaveInstructionsRequest,
    SaveTemplateDesignRequest
)
from app.execution.test_generator_excel import TestGeneratorExcelBuilder

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/generator", tags=["Test Case & Data Generator"])


@router.post("/parse-document")
async def parse_document(
    file: UploadFile = File(...),
    user: User = Depends(get_authenticated_user)
):
    """
    Ingests and parses uploaded requirements document (.pdf, .docx, .pptx, .xlsx, .csv, .txt).
    Extracts high-quality plain text for injection into the master prompt.
    """
    if not SystemKillSwitchManager.is_allowed("document_upload"):
        raise HTTPException(
            status_code=503,
            detail="Document upload is currently disabled by system administrator (Kill Switch Active)."
        )

    try:
        content = await file.read()
        parsed = extract_text_from_file(content, file.filename)
        return {
            "status": "SUCCESS",
            "filename": file.filename,
            "char_count": parsed["char_count"],
            "word_count": parsed["word_count"],
            "text": parsed["text"],
            "meta": {k: v for k, v in parsed.items() if k not in ("text", "filename", "char_count", "word_count")}
        }
    except Exception as e:
        logger.error(f"Failed to parse document {file.filename}: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")


async def _resolve_model_and_key(db: AsyncSession, requested_model: Optional[str], requested_provider: Optional[str]):
    """
    Finds active provider configuration and API key from database.
    """
    stmt = select(AIProviderSetting).where(AIProviderSetting.is_enabled != "false")
    res = await db.execute(stmt)
    active_providers = res.scalars().all()

    target_setting = None
    target_provider = None
    target_model = None

    if requested_provider:
        for prov in active_providers:
            if prov.provider.lower() == requested_provider.lower():
                target_setting = prov
                target_provider = prov.provider.lower()
                break

    # If not found, try to find provider by requested model
    if not target_setting and requested_model:
        for prov in active_providers:
            if requested_model in (prov.selected_models or []):
                target_setting = prov
                target_provider = prov.provider.lower()
                target_model = requested_model
                break

    # Fallback to first available enabled provider with a key
    if not target_setting:
        for prov in active_providers:
            has_key = bool(prov.api_key_encrypted) or bool(prov.api_keys)
            if has_key:
                target_setting = prov
                target_provider = prov.provider.lower()
                break

    if not target_setting:
        raise HTTPException(
            status_code=400,
            detail="No AI Provider configured or enabled. Please configure Groq, OpenAI, or Gemini in Admin Settings."
        )

    # Get API key
    key = None
    if target_setting.api_key_encrypted:
        key = decrypt_secret(target_setting.api_key_encrypted)
    elif target_setting.api_keys and len(target_setting.api_keys) > 0:
        first_key_obj = target_setting.api_keys[0]
        enc = first_key_obj.get("key_encrypted")
        if enc:
            key = decrypt_secret(enc)

    if not key:
        raise HTTPException(
            status_code=400,
            detail=f"No valid API key found for provider '{target_provider}'. Please update API keys in Admin Settings."
        )

    # Model resolution
    if not target_model:
        if requested_model:
            target_model = requested_model
        elif target_setting.selected_models and len(target_setting.selected_models) > 0:
            target_model = target_setting.selected_models[0]
        else:
            if target_provider == "groq":
                target_model = "openai/gpt-oss-120b"
            elif target_provider == "openai":
                target_model = "gpt-4o-mini"
            elif target_provider == "gemini":
                target_model = "gemini-1.5-flash"
            else:
                target_model = "default"

    return target_provider, target_model, key


def _clean_json_markdown(text: str) -> str:
    """Removes markdown code block formatting to retrieve raw JSON."""
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    # Find start and end of JSON array or object
    start_bracket = cleaned.find("[")
    start_brace = cleaned.find("{")
    
    if start_bracket != -1 and (start_brace == -1 or start_bracket < start_brace):
        end_bracket = cleaned.rfind("]")
        if end_bracket != -1:
            return cleaned[start_bracket:end_bracket + 1]
    elif start_brace != -1:
        end_brace = cleaned.rfind("}")
        if end_brace != -1:
            return cleaned[start_brace:end_brace + 1]

    return cleaned


def validate_hierarchy_and_template(
    columns: List[ColumnConfig],
    entity_levels: Optional[List[EntityLevel]] = None,
    mode: str = "both"
) -> None:
    """
    Validates template structure and hierarchy to prevent:
    1. Empty column list
    2. Duplicate column names (prevents LLM JSON key overwrites and data loss)
    3. Empty entity levels (levels with 0 columns causing broken tables)
    4. Duplicate entity level names or IDs
    5. Combinatorial row explosions (> 100 leaf rows)
    6. Non-positive or out-of-range branching ratios
    """
    if not columns or len(columns) == 0:
        raise HTTPException(status_code=400, detail="Template must contain at least 1 column.")

    # 1. Check duplicate column names
    seen_col_names = set()
    for col in columns:
        clean_name = (col.name or "").strip().lower()
        if not clean_name:
            raise HTTPException(status_code=400, detail="Column name cannot be empty.")
        if clean_name in seen_col_names:
            raise HTTPException(
                status_code=400,
                detail=f"Duplicate column name detected: '{col.name}'. Column names must be unique to avoid data loss."
            )
        seen_col_names.add(clean_name)

    # 2. Check Entity Levels if configured and mode is 'both'
    if mode == "both" and entity_levels and len(entity_levels) > 0:
        seen_level_ids = set()
        seen_level_names = set()
        estimated_leaf_rows = 1

        for i, lvl in enumerate(entity_levels):
            clean_id = (lvl.id or "").strip()
            clean_lvl_name = (lvl.name or "").strip()

            if not clean_lvl_name:
                raise HTTPException(status_code=400, detail=f"Entity Level {i+1} must have a name.")

            if clean_id in seen_level_ids:
                raise HTTPException(status_code=400, detail=f"Duplicate entity level ID: '{lvl.id}'.")
            seen_level_ids.add(clean_id)

            if clean_lvl_name.lower() in seen_level_names:
                raise HTTPException(
                    status_code=400,
                    detail=f"Duplicate entity level name: '{lvl.name}'. Each hierarchy level must have a distinct name."
                )
            seen_level_names.add(clean_lvl_name.lower())

            # Check columns in level
            lvl_cols = lvl.columns if lvl.columns is not None else [c for c in columns if c.entity_id == lvl.id]
            if not lvl_cols:
                raise HTTPException(
                    status_code=400,
                    detail=f"Entity Level '{lvl.name}' (L{i+1}) has 0 columns. Every level must have at least 1 column or be removed."
                )

            # Check branching ratio bounds
            mult = lvl.max_items_per_parent
            if mult < 1 or mult > 15:
                raise HTTPException(
                    status_code=400,
                    detail=f"Level '{lvl.name}' has invalid branching ratio '{mult}'. Must be between 1 and 15."
                )
            estimated_leaf_rows *= mult

        # Check Combinatorial Explosion
        if estimated_leaf_rows > 100:
            raise HTTPException(
                status_code=400,
                detail=f"Combinatorial explosion detected: Multipliers will generate ~{estimated_leaf_rows} rows (maximum safe limit is 100). Please lower your 'Max / Parent' branching ratios."
            )


@router.post("/generate")
async def generate_test_suite(
    req: GenerateTestRequest,
    user: User = Depends(get_authenticated_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Generates test cases and/or test data strictly in structured JSON schema.
    AI only produces the content; Python handles all table layout and Excel generation.
    """
    if not SystemKillSwitchManager.is_allowed("ai_execution"):
        raise HTTPException(
            status_code=503,
            detail="AI Generation is currently disabled by system administrator (Kill Switch Active)."
        )

    # Validate hierarchy and template configuration
    validate_hierarchy_and_template(
        columns=req.columns,
        entity_levels=req.entity_levels,
        mode=req.mode
    )

    provider, model, api_key = await _resolve_model_and_key(db, req.model_id, req.provider)

    case_cols = [c for c in req.columns if c.scope == "case"]
    data_cols = [c for c in req.columns if c.scope == "data"]
    
    # If mode is flat, treat all columns in their given scope
    if req.mode == "test_case":
        case_cols = req.columns
        data_cols = []
    elif req.mode == "test_data":
        data_cols = req.columns
        case_cols = []

    has_n_levels = bool(req.mode == "both" and req.entity_levels and len(req.entity_levels) > 1)

    if has_n_levels and req.entity_levels:
        def build_recursive_example(levels: List[Any], idx: int = 0) -> Dict[str, Any]:
            if idx >= len(levels):
                return {}
            lvl = levels[idx]
            cols = lvl.columns or [c for c in req.columns if c.entity_id == lvl.id]
            f = {c.name: f"<{c.name} for {lvl.name}>" for c in cols}
            res = {"fields": f}
            if idx + 1 < len(levels):
                res["children"] = [build_recursive_example(levels, idx + 1)]
            return res

        schema_sample = [build_recursive_example(req.entity_levels)]
        level_rules = "\n".join([
            f"- Level {i+1} ({lvl.name}): generate up to {lvl.max_items_per_parent} items per parent with columns: {[c.name for c in (lvl.columns or [col for col in req.columns if col.entity_id == lvl.id])]}"
            for i, lvl in enumerate(req.entity_levels)
        ])

        json_schema_desc = f"""
Return ONLY a valid JSON array of objects representing the root entity '{req.entity_levels[0].name}'.
Structure MUST follow this EXACT recursive hierarchy:
{json.dumps(schema_sample, indent=2)}

Hierarchy constraints:
{level_rules}
- For each node, 'fields' MUST contain values for its designated columns.
- 'children' contains the list of child entities for the next level down.
"""
    elif req.mode == "both":
        sample_case_fields = {c.name: f"<{c.name} value>" for c in case_cols}
        sample_data_row = {c.name: f"<{c.name} value>" for c in data_cols}
        json_schema_desc = f"""
Return ONLY a valid JSON array of objects with this EXACT structure:
[
  {{
    "case_fields": {json.dumps(sample_case_fields, indent=6)},
    "data_rows": [
      {json.dumps(sample_data_row, indent=6)}
    ]
  }}
]
Rules:
- Generate up to {req.max_test_cases} test cases in total.
- Under each test case, generate between 1 and {req.max_test_data_per_case} realistic, distinct test data variations in "data_rows".
- Each object in "data_rows" MUST have all data columns: {json.dumps([c.name for c in data_cols])}.
- "case_fields" MUST have all case columns: {json.dumps([c.name for c in case_cols])}.
"""
    elif req.mode == "test_case":
        sample_case_fields = {c.name: f"<{c.name} value>" for c in case_cols}
        json_schema_desc = f"""
Return ONLY a valid JSON array of objects with this EXACT structure:
[
  {{
    "case_fields": {json.dumps(sample_case_fields, indent=6)}
  }}
]
Rules:
- Generate up to {req.max_test_cases} distinct test cases.
- "case_fields" MUST include all columns: {json.dumps([c.name for c in case_cols])}.
"""
    else:  # test_data only
        sample_data_row = {c.name: f"<{c.name} value>" for c in data_cols}
        json_schema_desc = f"""
Return ONLY a valid JSON array of objects with this EXACT structure:
[
  {{
    "case_fields": {json.dumps(sample_data_row, indent=6)}
  }}
]
Rules:
- Generate up to {req.max_test_cases * req.max_test_data_per_case} test data variation rows.
- Each row MUST include all columns: {json.dumps([c.name for c in data_cols])}.
"""

    system_prompt = f"""You are an elite QA Test Architect and Test Automation Specialist.
Your task is to analyze software requirements and generate comprehensive, professional test cases and test data.
Coverage must include:
1. Positive / Happy path scenarios
2. Negative / Validation error scenarios
3. Boundary value analysis & Edge cases
4. Security / Injection / Authorization constraints (where applicable)

CRITICAL INSTRUCTIONS:
- You must output PURE RAW JSON only. Do not wrap in markdown quotes if possible, or use standard ```json.
- Do NOT output any explanatory text before or after the JSON.
- Every column specified in the user's template MUST be present in the keys.
- Write realistic, production-ready inputs, realistic dummy emails, phone numbers, payloads, and clear assertions.
"""

    user_prompt_parts = [
        f"### REQUIREMENTS / MASTER PROMPT:\n{req.master_prompt.strip()}"
    ]

    if req.document_text and req.document_text.strip():
        user_prompt_parts.append(
            f"### ATTACHED SPECIFICATION DOCUMENT CONTEXT:\n{req.document_text.strip()[:25000]}"
        )

    if req.instructions and req.instructions.strip():
        user_prompt_parts.append(
            f"### ADDITIONAL CUSTOM INSTRUCTIONS:\n{req.instructions.strip()}"
        )

    user_prompt_parts.append(
        f"### REQUIRED JSON OUTPUT SCHEMA:\n{json_schema_desc}\n\nGenerate the JSON now:"
    )

    full_user_prompt = "\n\n".join(user_prompt_parts)

    timeout = httpx.Timeout(90.0, connect=15.0)
    start_time = time.perf_counter()
    raw_response_text = ""
    prompt_tokens = 0
    completion_tokens = 0

    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            if provider == "groq":
                endpoint = "https://api.groq.com/openai/v1/chat/completions"
                headers = {"Authorization": f"Bearer {api_key}"}
                payload = {
                    "model": model,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": full_user_prompt}
                    ],
                    "temperature": 0.2,
                    "max_tokens": 4096,
                    "response_format": {"type": "json_object"} if "llama-3" in model or "gpt" in model else None
                }
                # Remove None fields
                payload = {k: v for k, v in payload.items() if v is not None}
                resp = await client.post(endpoint, headers=headers, json=payload)
                if resp.status_code != 200:
                    raise RuntimeError(f"Groq API error ({resp.status_code}): {resp.text}")
                res_data = resp.json()
                raw_response_text = res_data.get("choices", [{}])[0].get("message", {}).get("content", "")
                usage = res_data.get("usage", {})
                prompt_tokens = usage.get("prompt_tokens", 0)
                completion_tokens = usage.get("completion_tokens", 0)

            elif provider == "openai":
                endpoint = "https://api.openai.com/v1/chat/completions"
                headers = {"Authorization": f"Bearer {api_key}"}
                lower_m = model.lower()
                is_reasoning = lower_m.startswith("o1") or lower_m.startswith("o3")
                payload = {
                    "model": model,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": full_user_prompt}
                    ]
                }
                if is_reasoning:
                    payload["max_completion_tokens"] = 6000
                else:
                    payload["max_tokens"] = 4096
                    payload["temperature"] = 0.2
                    payload["response_format"] = {"type": "json_object"}

                resp = await client.post(endpoint, headers=headers, json=payload)
                if resp.status_code != 200:
                    raise RuntimeError(f"OpenAI API error ({resp.status_code}): {resp.text}")
                res_data = resp.json()
                raw_response_text = res_data.get("choices", [{}])[0].get("message", {}).get("content", "")
                usage = res_data.get("usage", {})
                prompt_tokens = usage.get("prompt_tokens", 0)
                completion_tokens = usage.get("completion_tokens", 0)

            elif provider == "gemini":
                gemini_model = model.replace("models/", "")
                endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{gemini_model}:generateContent?key={api_key}"
                combined = f"{system_prompt}\n\n{full_user_prompt}"
                payload = {
                    "contents": [{"parts": [{"text": combined}]}],
                    "generationConfig": {
                        "maxOutputTokens": 6000,
                        "temperature": 0.2,
                        "responseMimeType": "application/json"
                    }
                }
                resp = await client.post(endpoint, json=payload)
                if resp.status_code != 200:
                    raise RuntimeError(f"Google Gemini error ({resp.status_code}): {resp.text}")
                res_data = resp.json()
                candidates = res_data.get("candidates", [])
                if candidates and "content" in candidates[0]:
                    parts = candidates[0]["content"].get("parts", [])
                    if parts:
                        raw_response_text = parts[0].get("text", "")
                usage_meta = res_data.get("usageMetadata", {})
                prompt_tokens = usage_meta.get("promptTokenCount", 0)
                completion_tokens = usage_meta.get("candidatesTokenCount", 0)
            else:
                raise HTTPException(status_code=400, detail=f"Unsupported provider '{provider}'")

    except Exception as ex:
        latency_ms = (time.perf_counter() - start_time) * 1000
        await record_llm_usage(
            db=db,
            user_id=user.id,
            provider=provider,
            model=model,
            prompt_tokens=prompt_tokens,
            completion_tokens=completion_tokens,
            latency_ms=latency_ms,
            request_type="TEST_GENERATION",
            status="FAILED",
            error=str(ex)
        )
        logger.error(f"Error calling LLM provider {provider}: {ex}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"AI generation failed: {str(ex)}")

    latency_ms = (time.perf_counter() - start_time) * 1000

    # Record successful token usage
    await record_llm_usage(
        db=db,
        user_id=user.id,
        provider=provider,
        model=model,
        prompt_tokens=prompt_tokens,
        completion_tokens=completion_tokens,
        latency_ms=latency_ms,
        request_type="TEST_GENERATION",
        status="SUCCESS"
    )

    # Clean & Parse JSON
    cleaned_json = _clean_json_markdown(raw_response_text)
    parsed_records = []

    try:
        data = json.loads(cleaned_json)
        # If response was wrapped in an object like {"test_cases": [...]}
        if isinstance(data, dict):
            for k in ("test_cases", "data", "cases", "items", "results"):
                if k in data and isinstance(data[k], list):
                    data = data[k]
                    break
            if isinstance(data, dict):
                # Flat single item
                data = [data]

        if isinstance(data, list):
            parsed_records = data
        else:
            parsed_records = []
    except Exception as parse_err:
        logger.warning(f"Initial JSON parse failed: {parse_err}. Attempting partial recovery...")
        # Emergency regex extraction of JSON objects
        matches = re.findall(r"\{[^{}]*\}", cleaned_json)
        for m in matches:
            try:
                parsed_records.append(json.loads(m))
            except Exception:
                pass

    # Normalize structure for frontend preview and export
    normalized_data = []
    total_data_rows = 0

    if has_n_levels:
        # Assign unique IDs recursively to nodes if missing
        def tag_node_ids(node_list: List[Any], prefix: str = "node") -> List[Any]:
            for i, n in enumerate(node_list):
                if isinstance(n, dict):
                    if "id" not in n:
                        n["id"] = f"{prefix}-{i+1}"
                    children = n.get("children") or []
                    if children:
                        tag_node_ids(children, prefix=f"{n['id']}")
            return node_list

        normalized_data = tag_node_ids(parsed_records)
        total_data_rows = sum(TestGeneratorExcelBuilder.get_leaf_count(item) for item in normalized_data)
    else:
        for idx, item in enumerate(parsed_records):
            if req.mode == "both":
                c_fields = item.get("case_fields") if isinstance(item.get("case_fields"), dict) else {}
                if not c_fields:
                    c_fields = {c.name: item.get(c.name, item.get(c.id, "")) for c in case_cols}

                d_rows = item.get("data_rows") if isinstance(item.get("data_rows"), list) else []
                if not d_rows:
                    d_row = {c.name: item.get(c.name, item.get(c.id, "")) for c in data_cols}
                    d_rows = [d_row]

                total_data_rows += len(d_rows)
                normalized_data.append({
                    "id": f"tc-{idx+1}",
                    "case_fields": c_fields,
                    "data_rows": d_rows
                })
            else:
                c_fields = item.get("case_fields") if isinstance(item.get("case_fields"), dict) else item
                normalized_data.append({
                    "id": f"rec-{idx+1}",
                    "case_fields": c_fields,
                    "data_rows": []
                })
                total_data_rows += 1

    return {
        "status": "SUCCESS",
        "mode": req.mode,
        "provider": provider,
        "model": model,
        "latency_ms": round(latency_ms, 1),
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "total_tokens": prompt_tokens + completion_tokens,
        "total_cases": len(normalized_data),
        "total_data_rows": total_data_rows,
        "columns": [c.model_dump() for c in req.columns],
        "data": normalized_data
    }


@router.post("/export-excel")
async def export_excel_matrix(
    req: ExportExcelRequest,
    user: User = Depends(get_authenticated_user)
):
    """
    Generates downloadable Excel file via pure openpyxl with dynamic row merging and corporate styling.
    AI generates content; Python creates the production-grade spreadsheet.
    """
    try:
        excel_stream = TestGeneratorExcelBuilder.generate_excel(
            columns=req.columns,
            data=req.data,
            entity_levels=req.entity_levels,
            mode=req.mode,
            title=req.sheet_name or "AI Test Suite & Test Data Matrix",
            sheet_name=req.sheet_name or "Test Matrix"
        )
        safe_filename = req.filename or "Generated_Test_Cases_Matrix.xlsx"
        if not safe_filename.endswith(".xlsx"):
            safe_filename += ".xlsx"

        return StreamingResponse(
            excel_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        logger.error(f"Excel generation failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {str(e)}")


@router.post("/save-dataset")
async def save_generated_as_dataset(
    req: SaveDatasetRequest,
    user: User = Depends(get_authenticated_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Saves the generated test cases/data directly into the platform's Test Datasets repository
    for immediate execution in Flow Builder, Matrix Runner, or Swarm.
    """
    try:
        # Build headers list
        headers = [col.name for col in req.columns]
        rows = []

        def extract_leaf_rows(node_list: List[Dict[str, Any]], parent_vals: Dict[str, Any] = None) -> List[Dict[str, Any]]:
            if parent_vals is None:
                parent_vals = {}
            res_rows = []
            for item in node_list:
                node_fields = dict(parent_vals)
                extracted = item.get("fields") or item.get("case_fields") or {}
                if not extracted and isinstance(item, dict):
                    extracted = {k: v for k, v in item.items() if k not in ("children", "data_rows", "id", "case_fields", "fields")}
                node_fields.update(extracted)

                children = item.get("children") or item.get("data_rows") or []
                if not children:
                    for k in ("test_cases", "test_data", "data", "followups", "followup_prompts", "items"):
                        if k in item and isinstance(item[k], list) and len(item[k]) > 0:
                            children = item[k]
                            break

                if children:
                    res_rows.extend(extract_leaf_rows(children, node_fields))
                else:
                    res_rows.append(node_fields)
            return res_rows

        leaf_records = extract_leaf_rows(req.data)
        for leaf in leaf_records:
            row_vals = []
            for col in req.columns:
                val = leaf.get(col.name)
                if val is None:
                    val = leaf.get(col.id, "")
                row_vals.append(str(val) if val is not None else "")
            rows.append(row_vals)

        dataset = TestDataset(
            project_id=req.project_id,
            name=req.name,
            description=req.description or "Generated via AI Test Case & Data Studio",
            headers=headers,
            rows=rows
        )
        db.add(dataset)
        await db.commit()
        await db.refresh(dataset)

        return {
            "status": "SUCCESS",
            "message": "Dataset saved successfully",
            "dataset_id": dataset.id,
            "name": dataset.name,
            "total_rows": len(rows),
            "headers": headers
        }
    except Exception as e:
        logger.error(f"Failed to save generated dataset: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to save dataset: {str(e)}")


# =========================================================================
# PROJECT-LEVEL PERSISTENCE (PROMPT, INSTRUCTIONS, TEMPLATE DESIGN)
# =========================================================================

@router.get("/projects/{project_id}/config")
async def get_project_generator_config(
    project_id: str,
    user: User = Depends(get_authenticated_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Retrieves the saved prompt, instructions, and Excel template design for a specific project.
    """
    stmt = select(Project).where(Project.id == project_id)
    res = await db.execute(stmt)
    project = res.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    tmpl = dict(project.report_template or {})
    gen_settings = tmpl.get("generator_settings", {})

    return {
        "status": "SUCCESS",
        "project_id": project_id,
        "master_prompt": gen_settings.get("master_prompt", ""),
        "instructions": gen_settings.get("instructions", ""),
        "template_design": gen_settings.get("template_design", None)
    }


@router.post("/projects/{project_id}/save-prompt")
async def save_project_master_prompt(
    project_id: str,
    req: SavePromptRequest,
    user: User = Depends(get_authenticated_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Saves master prompt requirement on the project level.
    """
    stmt = select(Project).where(Project.id == project_id)
    res = await db.execute(stmt)
    project = res.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    tmpl = dict(project.report_template or {})
    if "generator_settings" not in tmpl:
        tmpl["generator_settings"] = {}
    tmpl["generator_settings"]["master_prompt"] = req.prompt
    project.report_template = tmpl
    flag_modified(project, "report_template")
    await db.commit()

    return {
        "status": "SUCCESS",
        "message": "Master prompt saved to project successfully",
        "project_id": project_id
    }


@router.post("/projects/{project_id}/save-instructions")
async def save_project_instructions(
    project_id: str,
    req: SaveInstructionsRequest,
    user: User = Depends(get_authenticated_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Saves testing instructions and guidelines on the project level.
    """
    stmt = select(Project).where(Project.id == project_id)
    res = await db.execute(stmt)
    project = res.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    tmpl = dict(project.report_template or {})
    if "generator_settings" not in tmpl:
        tmpl["generator_settings"] = {}
    tmpl["generator_settings"]["instructions"] = req.instructions
    project.report_template = tmpl
    flag_modified(project, "report_template")
    await db.commit()

    return {
        "status": "SUCCESS",
        "message": "Instructions saved to project successfully",
        "project_id": project_id
    }


@router.post("/projects/{project_id}/save-template")
async def save_project_template_design(
    project_id: str,
    req: SaveTemplateDesignRequest,
    user: User = Depends(get_authenticated_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Saves Excel template design (dynamic columns, merge rules, and scopes) on the project level.
    """
    # Validate hierarchy and template configuration
    validate_hierarchy_and_template(
        columns=req.columns,
        entity_levels=req.entity_levels,
        mode=req.mode or "both"
    )

    stmt = select(Project).where(Project.id == project_id)
    res = await db.execute(stmt)
    project = res.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    tmpl = dict(project.report_template or {})
    if "generator_settings" not in tmpl:
        tmpl["generator_settings"] = {}
    tmpl["generator_settings"]["template_design"] = req.model_dump()
    project.report_template = tmpl
    flag_modified(project, "report_template")
    await db.commit()

    return {
        "status": "SUCCESS",
        "message": "Excel template design saved to project successfully",
        "project_id": project_id,
        "template_design": tmpl["generator_settings"]["template_design"]
    }


@router.post("/import-template-excel")
async def import_template_excel(
    file: UploadFile = File(...),
    user: User = Depends(get_authenticated_user)
):
    """
    Parses an uploaded Excel (.xlsx) or CSV template to extract table headers
    and construct a clean dynamic entity hierarchy and column configuration.
    """
    if not file.filename.lower().endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel (.xlsx) or CSV file.")

    content = await file.read()
    raw_columns = []
    
    if file.filename.lower().endswith(".csv"):
        import csv
        text_stream = io.StringIO(content.decode("utf-8", errors="ignore"))
        reader = csv.reader(text_stream)
        for row in reader:
            non_empty = [c.strip() for c in row if c and c.strip()]
            if len(non_empty) >= 1:
                raw_columns = non_empty
                break
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            best_row = 1
            max_headers = 0

            for r in range(1, min(15, ws.max_row + 1)):
                vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                str_vals = [str(v).strip() for v in vals if v is not None and str(v).strip() != ""]
                # Skip title / meta banner rows
                if len(str_vals) > max_headers and not any("GENERATED BY" in s.upper() for s in str_vals):
                    max_headers = len(str_vals)
                    best_row = r

            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=best_row, column=c).value
                if val is not None and str(val).strip() != "":
                    raw_columns.append(str(val).strip())
        except Exception as e:
            logger.error(f"Failed to parse Excel template {file.filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    if not raw_columns:
        raise HTTPException(status_code=400, detail="No column headers could be found in the uploaded file.")

    # Deduplicate column names if necessary
    seen = {}
    clean_cols = []
    for c in raw_columns:
        base = c
        if base.lower() in seen:
            seen[base.lower()] += 1
            clean_cols.append(f"{base}_{seen[base.lower()]}")
        else:
            seen[base.lower()] = 0
            clean_cols.append(base)

    # Construct 2-level hierarchy (Parent and Child)
    split_idx = max(1, len(clean_cols) // 2) if len(clean_cols) > 2 else 1
    lvl1_names = clean_cols[:split_idx]
    lvl2_names = clean_cols[split_idx:]

    lvl1_id = "lvl_imported_parent"
    lvl2_id = "lvl_imported_child"

    lvl1_cols = [
        ColumnConfig(
            id=f"col_imp_{i+1}",
            name=name,
            scope=lvl1_id,
            entity_id=lvl1_id,
            merge_rows=True,
            description=f"Imported field: {name}"
        )
        for i, name in enumerate(lvl1_names)
    ]

    lvl2_cols = [
        ColumnConfig(
            id=f"col_imp_{i + 1 + len(lvl1_names)}",
            name=name,
            scope=lvl2_id,
            entity_id=lvl2_id,
            merge_rows=False,
            description=f"Imported field: {name}"
        )
        for i, name in enumerate(lvl2_names)
    ]

    entity_levels = [
        EntityLevel(
            id=lvl1_id,
            name="Imported Test Scenario",
            description="Parent entity level imported from Excel",
            max_items_per_parent=3,
            columns=lvl1_cols
        ),
        EntityLevel(
            id=lvl2_id,
            name="Imported Test Data & Outcomes",
            description="Child entity level imported from Excel",
            max_items_per_parent=3,
            columns=lvl2_cols
        )
    ]

    all_columns = lvl1_cols + lvl2_cols

    return {
        "status": "SUCCESS",
        "filename": file.filename,
        "mode": "both",
        "columns": [c.model_dump() for c in all_columns],
        "entity_levels": [l.model_dump() for l in entity_levels]
    }

