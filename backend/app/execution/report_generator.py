import io
import json
import re
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    """
    Production-grade Excel (.xlsx) test execution report generator.
    Produces richly styled multi-tab spreadsheets with:
    - Executive Summary Dashboard (KPIs, Pass Rate %, Average Duration)
    - Detailed Scenarios Data Table with:
      - 1 column for Initial Message & 1 column for Initial Response
      - 1 column for Follow-up Question & 1 column for Follow-up Response
      - Automatic row-wise expansion for multi-turn followup questions & responses
      - Full handling for 0 follow-up questions or workflows without follow-up nodes.
    """

    # Color definitions (Hex)
    HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Slate 800 Navy
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    SUMMARY_HDR_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")  # Teal 700
    SUMMARY_HDR_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Emerald 100
    PASS_FONT = Font(name="Calibri", size=10, bold=True, color="166534")  # Emerald 800

    FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Rose 100
    FAIL_FONT = Font(name="Calibri", size=10, bold=True, color="991B1B")  # Rose 800

    RUNNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber 100
    RUNNING_FONT = Font(name="Calibri", size=10, bold=True, color="92400E")  # Amber 800

    HYPERLINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")  # Standard Excel Link

    CARD_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    BORDER_THIN = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    DEFAULT_COLUMNS = [
        {"id": "scenario_index", "label": "Scenario #", "enabled": True},
        {"id": "title", "label": "Scenario Title / Prompt", "enabled": True},
        {"id": "input_message", "label": "Initial Message", "enabled": True},
        {"id": "response_message", "label": "Initial Response", "enabled": True},
        {"id": "input_followup", "label": "Follow-up Question", "enabled": True},
        {"id": "response_followup", "label": "Follow-up Response", "enabled": True},
        {"id": "chat_url", "label": "Chat Session URL", "enabled": True},
        {"id": "status", "label": "Execution Status", "enabled": True},
        {"id": "duration_ms", "label": "Latency (ms)", "enabled": True},
        {"id": "started_at", "label": "Executed At", "enabled": True}
    ]

    @classmethod
    def generate_report(
        cls,
        project_name: str,
        executions: List[Any],
        template_config: Optional[Dict[str, Any]] = None
    ) -> io.BytesIO:
        wb = openpyxl.Workbook()
        # Default active sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_data = wb.create_sheet(title="Detailed Scenarios")

        template = template_config or {}
        columns = template.get("columns") or cls.DEFAULT_COLUMNS
        active_cols = [c for c in columns if c.get("enabled", True)]
        if not active_cols:
            active_cols = cls.DEFAULT_COLUMNS

        include_summary = template.get("include_summary", True)

        # 1. Generate Detailed Scenarios Sheet (with Row-wise Follow-up Expansion)
        cls._build_scenarios_sheet(ws_data, executions, active_cols, template)

        # 2. Generate Summary Dashboard Sheet
        if include_summary:
            cls._build_summary_sheet(ws_summary, project_name, executions, template)
        else:
            wb.remove(ws_summary)

        # Save to in-memory bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def _build_scenarios_sheet(
        cls,
        ws: Any,
        executions: List[Any],
        columns: List[Dict[str, Any]],
        template: Dict[str, Any]
    ):
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True

        # Header Row
        headers = [col.get("label") or col.get("id") for col in columns]
        ws.append(headers)

        header_row = ws[1]
        ws.row_dimensions[1].height = 26

        for cell in header_row:
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cls.BORDER_THIN

        # Freeze header row
        ws.freeze_panes = "A2"

        # Track max length per column for auto-sizing
        col_widths = {i + 1: len(str(h)) + 4 for i, h in enumerate(headers)}

        current_excel_row = 2
        merge_scenario_cells = template.get("merge_scenario_cells", True)
        format_urls_hyperlink = template.get("format_urls_hyperlink", True)

        # Populate rows with automatic multi-turn row expansion & dynamic cell merging
        for exec_idx, exec_item in enumerate(executions):
            row_data = cls._extract_row_data(exec_item, exec_idx)
            followup_turns = row_data.get("followup_turns", [])
            num_rows = max(1, len(followup_turns))
            start_row = current_excel_row
            end_row = current_excel_row + num_rows - 1

            for turn_i in range(num_rows):
                ws.row_dimensions[current_excel_row].height = 24

                row_values = []
                for col_idx, col in enumerate(columns, start=1):
                    col_id = col.get("id", "")
                    val = cls._resolve_turn_column_value(col_id, row_data, exec_item, turn_i, num_rows)

                    # Format strings if dictionary or list, preserve numbers
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        cell_val = val
                        cell_str = str(val)
                    elif isinstance(val, (dict, list)):
                        cell_val = cls._clean_response_payload(val)
                        cell_str = str(cell_val)
                    else:
                        cell_val = str(val) if val is not None else ""
                        cell_str = cell_val

                    row_values.append(cell_val)

                    # Update column width tracker
                    line_lengths = [len(line) for line in cell_str.split("\n")]
                    max_line_len = max(line_lengths) if line_lengths else len(cell_str)
                    col_widths[col_idx] = max(col_widths[col_idx], min(max_line_len + 3, 60))

                ws.append(row_values)

                # Style each cell in row
                current_row = ws[current_excel_row]
                for col_idx, cell in enumerate(current_row, start=1):
                    cell.border = cls.BORDER_THIN
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                    cell_str_val = str(cell.value or "").strip()
                    col_id = columns[col_idx - 1].get("id", "")

                    # Apply Clickable Hyperlink formatting for URLs
                    if format_urls_hyperlink and cell_str_val.startswith(("http://", "https://")):
                        cell.hyperlink = cell_str_val
                        cell.font = cls.HYPERLINK_FONT

                    # Apply status highlight
                    if col_id in ("status", "execution_status") and cell_str_val:
                        status_upper = cell_str_val.upper()
                        if "PASS" in status_upper or "SUCCESS" in status_upper:
                            cell.fill = cls.PASS_FILL
                            cell.font = cls.PASS_FONT
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif "FAIL" in status_upper or "ERROR" in status_upper:
                            cell.fill = cls.FAIL_FILL
                            cell.font = cls.FAIL_FONT
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif "RUN" in status_upper:
                            cell.fill = cls.RUNNING_FILL
                            cell.font = cls.RUNNING_FONT
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Center align index or numbers
                    if col_id in ("scenario_index", "total_turns", "quality_score", "safety_score"):
                        cell.alignment = Alignment(horizontal="center", vertical="top")

                current_excel_row += 1

            # Apply Dynamic Cell Merging across multi-row scenarios
            if num_rows > 1 and merge_scenario_cells:
                for col_idx, col in enumerate(columns, start=1):
                    col_id = col.get("id", "").lower().strip()
                    merge_rule = col.get("merge_rule", "by_scenario")

                    should_merge = False
                    if merge_rule in ("by_scenario", "scenario", "="):
                        if col_id not in ("input_followup", "response_followup", "followup", "follow_up_questions_response"):
                            should_merge = True
                    elif merge_rule in ("same_value", "value", "=="):
                        first_val = ws.cell(row=start_row, column=col_idx).value
                        if first_val and all(ws.cell(row=r, column=col_idx).value in (first_val, "", None) for r in range(start_row, end_row + 1)):
                            should_merge = True
                    elif merge_rule == "none":
                        should_merge = False

                    if should_merge:
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col_idx,
                            end_row=end_row,
                            end_column=col_idx
                        )
                        top_cell = ws.cell(row=start_row, column=col_idx)
                        for r in range(start_row, end_row + 1):
                            c = ws.cell(row=r, column=col_idx)
                            c.border = cls.BORDER_THIN

                        top_str_val = str(top_cell.value or "").strip()
                        if format_urls_hyperlink and top_str_val.startswith(("http://", "https://")):
                            top_cell.hyperlink = top_str_val
                            top_cell.font = cls.HYPERLINK_FONT

                        if col_id in ("scenario_index", "status", "execution_status"):
                            top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        else:
                            top_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Auto-fit column widths
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width, 14)

        # Enable Auto-Filter
        if current_excel_row > 2 and len(columns) > 0:
            last_col_letter = get_column_letter(len(columns))
            ws.auto_filter.ref = f"A1:{last_col_letter}{current_excel_row - 1}"

    @staticmethod
    def _is_passed(status_val: Any) -> bool:
        if hasattr(status_val, "value"):
            s = str(status_val.value).upper()
        else:
            s = str(status_val or "").upper()
        return "PASS" in s or "SUCCESS" in s

    @staticmethod
    def _is_failed(status_val: Any) -> bool:
        if hasattr(status_val, "value"):
            s = str(status_val.value).upper()
        else:
            s = str(status_val or "").upper()
        return "FAIL" in s or "ERROR" in s

    @classmethod
    def _build_summary_sheet(
        cls,
        ws: Any,
        project_name: str,
        executions: List[Any],
        template: Dict[str, Any]
    ):
        ws.views.sheetView[0].showGridLines = True

        # Calculate KPIs
        total_runs = len(executions)
        passed_count = sum(1 for e in executions if cls._is_passed(getattr(e, "status", "")))
        failed_count = sum(1 for e in executions if cls._is_failed(getattr(e, "status", "")))
        pass_rate = round((passed_count / total_runs * 100.0), 1) if total_runs > 0 else 0.0

        durations = [getattr(e, "total_duration_ms", 0) or 0 for e in executions]
        avg_dur_ms = round(sum(durations) / total_runs, 1) if total_runs > 0 else 0.0

        total_tokens = sum((getattr(e, "total_tokens", 0) or 0) for e in executions)

        # Title Block
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"📊 AI Agent Automation Test Report — {project_name}"
        title_cell.fill = cls.SUMMARY_HDR_FILL
        title_cell.font = cls.SUMMARY_HDR_FONT
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 36

        # Subtitle metadata
        ws["A2"] = f"Generated On: {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M:%S UTC')}  •  Total Scenarios Evaluated: {total_runs}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")
        ws.row_dimensions[2].height = 20

        # KPI Header Cards (Row 4 & 5)
        kpi_defs = [
            ("Total Scenarios", f"{total_runs}", "1E293B", "F8FAFC"),
            ("Passed", f"{passed_count}", "166534", "DCFCE7"),
            ("Failed", f"{failed_count}", "991B1B", "FEE2E2"),
            ("Pass Rate", f"{pass_rate}%", "0F766E", "CCFBF1"),
            ("Avg Latency", f"{avg_dur_ms} ms", "334155", "F1F5F9"),
            ("Total Tokens", f"{total_tokens:,}", "4338CA", "EEF2FF"),
        ]

        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 26

        for i, (label, val, text_color, bg_color) in enumerate(kpi_defs, start=1):
            col_letter = get_column_letter(i)
            # Label
            c_lbl = ws[f"{col_letter}4"]
            c_lbl.value = label.upper()
            c_lbl.font = Font(name="Calibri", size=9, bold=True, color="64748B")
            c_lbl.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            c_lbl.alignment = Alignment(horizontal="center", vertical="center")
            c_lbl.border = cls.BORDER_THIN

            # Value
            c_val = ws[f"{col_letter}5"]
            c_val.value = val
            c_val.font = Font(name="Calibri", size=14, bold=True, color=text_color)
            c_val.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            c_val.alignment = Alignment(horizontal="center", vertical="center")
            c_val.border = cls.BORDER_THIN

        # Execution Runs Overview Table (Starting at Row 8)
        ws["A7"] = "EXECUTION RUNS SUMMARY"
        ws["A7"].font = Font(name="Calibri", size=11, bold=True, color="1E293B")
        ws.row_dimensions[7].height = 22

        table_headers = ["Scenario #", "Title / Prompt", "Status", "Duration (ms)", "Turns", "Correlation ID"]
        ws.append(table_headers)
        ws.row_dimensions[8].height = 24

        hdr_row = ws[8]
        for cell in hdr_row:
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cls.BORDER_THIN

        for r_idx, exec_item in enumerate(executions, start=9):
            ws.row_dimensions[r_idx].height = 20
            row_data = cls._extract_row_data(exec_item, r_idx - 9)

            sc_num = row_data.get("scenario_index", r_idx - 8)
            sc_title = row_data.get("title", "")
            sc_status = row_data.get("status", "PASSED")
            sc_dur = row_data.get("duration_ms", 0)
            sc_turns = row_data.get("total_turns", 1)
            sc_corr = row_data.get("correlation_id", "")

            ws.append([sc_num, sc_title, sc_status, sc_dur, sc_turns, sc_corr])

            curr_row = ws[r_idx]
            for col_i, cell in enumerate(curr_row, start=1):
                cell.border = cls.BORDER_THIN
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center")

                if col_i in (1, 4, 5):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_i == 3:
                    if "PASS" in str(cell.value).upper():
                        cell.fill = cls.PASS_FILL
                        cell.font = cls.PASS_FONT
                    else:
                        cell.fill = cls.FAIL_FILL
                        cell.font = cls.FAIL_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 30

    @classmethod
    def _extract_row_data(cls, exec_item: Any, default_idx: int) -> Dict[str, Any]:
        """Extracts normalized scenario metadata, initial query/response, and all multi-turn followup questions and responses."""
        corr = getattr(exec_item, "correlation_id", "") or ""
        match = re.search(r'-s(\d+)-', corr, re.IGNORECASE) or re.search(r'-r(\d+)-', corr, re.IGNORECASE) or re.search(r'row-(\d+)', corr, re.IGNORECASE) or re.search(r'Scenario-(\d+)', corr, re.IGNORECASE)
        
        ctx = getattr(exec_item, "runtime_context", {}) or {}
        scenario_idx = ctx.get("scenario_index") if isinstance(ctx, dict) else None
        if scenario_idx is None:
            scenario_idx = int(match.group(1)) if match else (getattr(exec_item, "dataset_row_index", None) or default_idx) + 1

        dataset_vars = ctx.get("dataset_vars", {}) if isinstance(ctx, dict) else {}
        turns = ctx.get("turns", []) if isinstance(ctx, dict) else []

        scenario_title = (ctx.get("scenario") if isinstance(ctx, dict) else None) or dataset_vars.get("message") or dataset_vars.get("query") or f"Scenario #{scenario_idx}"

        # Status
        st = getattr(exec_item, "status", "PASSED")
        status_str = st.value if hasattr(st, "value") else str(st)

        # Output / Captured variables
        captured_vars: Dict[str, Any] = {}
        if isinstance(dataset_vars, dict):
            captured_vars.update(dataset_vars)

        steps = getattr(exec_item, "steps", []) or []
        traces = getattr(exec_item, "trace_events", []) or []

        # 1. Initial Message / Query Extraction across all sources
        initial_query = ""
        for k, v in dataset_vars.items():
            if k.lower() in ("message", "query", "prompt", "user_query", "input_message", "input", "question", "prompt_message") and v:
                initial_query = str(v)
                break

        if not initial_query and isinstance(ctx, dict):
            initial_query = str(ctx.get("scenario") or ctx.get("prompt") or ctx.get("title") or ctx.get("message") or "")

        if not initial_query and isinstance(turns, list) and len(turns) > 0:
            t0 = turns[0]
            if isinstance(t0, dict):
                initial_query = str(t0.get("message") or t0.get("query") or t0.get("prompt") or t0.get("scenario") or "")

        if not initial_query and steps:
            for s in steps:
                s_inp = getattr(s, "input_data", {}) or {}
                if isinstance(s_inp, dict):
                    q = s_inp.get("message") or s_inp.get("query") or s_inp.get("prompt")
                    if q:
                        initial_query = str(q)
                        break

        if not initial_query and traces:
            for t in traces:
                t_raw = getattr(t, "raw_payload", {}) or {}
                if isinstance(t_raw, dict):
                    q = t_raw.get("message") or t_raw.get("query") or t_raw.get("prompt")
                    if q:
                        initial_query = str(q)
                        break

        scenario_title = initial_query or (ctx.get("scenario") if isinstance(ctx, dict) else None) or f"Scenario #{scenario_idx}"

        # Collect captured variables from steps & traces
        for s in steps:
            s_out = getattr(s, "output_data", None)
            s_key = getattr(s, "node_key", "") or ""
            if s_out is not None:
                if isinstance(s_out, dict) and "captured_variables" in s_out:
                    captured_vars.update(s_out["captured_variables"])
                elif isinstance(s_out, dict) and "response" in s_out:
                    captured_vars[s_key] = s_out["response"]
                else:
                    captured_vars[s_key] = s_out

        for t in traces:
            t_raw = getattr(t, "raw_payload", None)
            if isinstance(t_raw, dict) and "captured_variables" in t_raw:
                captured_vars.update(t_raw["captured_variables"])

        initial_response = captured_vars.get("message_api_response") or captured_vars.get("initial_response") or captured_vars.get("response") or ""

        # 2. Extract All Follow-up Questions (handling multiple questions per scenario)
        followup_queries: List[str] = []
        if isinstance(turns, list) and len(turns) > 0:
            for t_idx, t in enumerate(turns):
                if isinstance(t, dict):
                    fq = t.get("followup") or t.get("follow_up")
                    if fq and str(fq) not in followup_queries:
                        followup_queries.append(str(fq))
                    elif t_idx > 0:
                        q = t.get("query") or t.get("message")
                        if q and str(q) not in followup_queries:
                            followup_queries.append(str(q))

        if not followup_queries:
            # Check dataset_vars for followup, followup_1, followup_2, etc.
            for k, v in dataset_vars.items():
                if "followup" in k.lower() and v and str(v) not in followup_queries:
                    followup_queries.append(str(v))

        # 3. Extract All Follow-up Responses (handling multiple turns & 0 turns)
        followup_responses: List[str] = []

        # From turn steps (e.g. node-0403_turn_1, node-0403_turn_2)
        for s in steps:
            s_key = getattr(s, "node_key", "") or ""
            s_out = getattr(s, "output_data", None)
            if s_out is not None:
                turn_match = re.search(r'_turn_(\d+)', s_key, re.IGNORECASE)
                if turn_match:
                    t_idx = int(turn_match.group(1)) - 1
                    cleaned = cls._clean_response_payload(s_out)
                    while len(followup_responses) <= t_idx:
                        followup_responses.append("")
                    followup_responses[t_idx] = str(cleaned)

        # From turn traces (e.g. "Follow-up Questions (Turn #1)")
        for tr in traces:
            tr_title = getattr(tr, "title", "") or ""
            tr_raw = getattr(tr, "raw_payload", None)
            if tr_raw is not None:
                turn_match = re.search(r'Turn\s*#?(\d+)', tr_title, re.IGNORECASE)
                if turn_match and any(w in tr_title.lower() for w in ("follow", "question", "turn")):
                    t_idx = int(turn_match.group(1)) - 1
                    cleaned = cls._clean_response_payload(tr_raw)
                    while len(followup_responses) <= t_idx:
                        followup_responses.append("")
                    if not followup_responses[t_idx]:
                        followup_responses[t_idx] = str(cleaned)

        # Fallback from captured_vars (e.g. follow_up_questions_response (Turn #1) or follow_up_questions_response)
        for k, v in captured_vars.items():
            turn_match = re.search(r'Turn\s*#?(\d+)', k, re.IGNORECASE)
            if turn_match and "follow" in k.lower():
                t_idx = int(turn_match.group(1)) - 1
                cleaned = cls._clean_response_payload(v)
                while len(followup_responses) <= t_idx:
                    followup_responses.append("")
                if not followup_responses[t_idx]:
                    followup_responses[t_idx] = str(cleaned)

        if not followup_responses and "follow_up_questions_response" in captured_vars:
            resp_val = captured_vars["follow_up_questions_response"]
            if resp_val:
                followup_responses.append(str(cls._clean_response_payload(resp_val)))

        # 4. Build Paired Follow-up Turns (0 questions handled cleanly as empty list)
        followup_turns: List[Dict[str, str]] = []
        max_f = max(len(followup_queries), len(followup_responses))
        for i in range(max_f):
            q = followup_queries[i] if i < len(followup_queries) else ""
            r = followup_responses[i] if i < len(followup_responses) else ""
            if q or r:
                followup_turns.append({"query": q, "response": r})

        total_turns = max(len(followup_turns) + 1 if followup_turns else 1, len(turns) if turns else 1)

        # 5. Extract Chat URL from CHAT_URL_CREATOR node, captured variables, or steps
        chat_url = captured_vars.get("chat_url") or captured_vars.get("url") or dataset_vars.get("chat_url") or ""
        if not chat_url and steps:
            for s in steps:
                s_type = getattr(s, "node_type", "") or ""
                s_out = getattr(s, "output_data", {}) or {}
                if isinstance(s_out, dict):
                    if s_type == "CHAT_URL_CREATOR" or "chat_url" in s_out or "url" in s_out:
                        chat_url = s_out.get("chat_url") or s_out.get("url") or ""
                        if chat_url:
                            break

        return {
            "scenario_index": scenario_idx,
            "title": scenario_title,
            "status": status_str,
            "duration_ms": round(getattr(exec_item, "total_duration_ms", 0.0) or 0.0, 1),
            "started_at": str(getattr(exec_item, "started_at", "") or getattr(exec_item, "created_at", "") or ""),
            "completed_at": str(getattr(exec_item, "completed_at", "") or ""),
            "correlation_id": corr,
            "total_turns": total_turns,
            "quality_score": getattr(exec_item, "quality_score", None),
            "safety_score": getattr(exec_item, "safety_score", None),
            "total_tokens": getattr(exec_item, "total_tokens", 0) or 0,
            "error_message": getattr(exec_item, "error_message", None),
            "initial_query": initial_query,
            "initial_response": cls._clean_response_payload(initial_response),
            "chat_url": str(chat_url),
            "followup_turns": followup_turns,
            "dataset_vars": dataset_vars,
            "captured_vars": captured_vars,
        }

    @classmethod
    def _resolve_turn_column_value(
        cls,
        col_id: str,
        row_data: Dict[str, Any],
        exec_item: Any,
        turn_idx: int,
        total_rows: int
    ) -> Any:
        """Resolves column value for specific multi-turn row index (turn_idx = 0 is main scenario row)."""
        c_lower = col_id.lower().strip()
        dvars = row_data.get("dataset_vars", {})
        cvars = row_data.get("captured_vars", {})
        followup_turns = row_data.get("followup_turns", [])

        # 1. Follow-up Question Column (Row-wise dynamically populated)
        if c_lower in ("followup", "input_followup", "follow_up", "followup_question", "follow_up_question", "followup_query"):
            if followup_turns and turn_idx < len(followup_turns):
                return followup_turns[turn_idx].get("query", "")
            return ""

        # 2. Follow-up Response Column (Row-wise dynamically populated)
        if c_lower in ("response_followup", "follow_up_questions_response", "followup_response", "follow_up_response"):
            if followup_turns and turn_idx < len(followup_turns):
                return followup_turns[turn_idx].get("response", "")
            return ""

        # 3. Scenario Index
        if c_lower in ("scenario_index", "scenario_num", "index", "#"):
            return row_data.get("scenario_index")

        # 4. Status
        if c_lower in ("status", "execution_status"):
            return row_data.get("status")

        # 5. Scenario Title & Initial Query (Shown on first row of scenario)
        if c_lower in ("title", "scenario", "scenario_title"):
            return row_data.get("title") if turn_idx == 0 else ""

        if c_lower in ("input_message", "message", "query", "prompt", "initial_message", "initial_query"):
            return (row_data.get("initial_query") or dvars.get("message") or dvars.get("query") or "") if turn_idx == 0 else ""

        # 6. Initial Response (Shown on first row of scenario)
        if c_lower in ("response_message", "message_api_response", "initial_response", "answer"):
            return row_data.get("initial_response") if turn_idx == 0 else ""

        # 7. Latency, Timestamps & Metrics (Shown on first row of scenario)
        if c_lower in ("duration_ms", "duration", "latency"):
            return row_data.get("duration_ms") if turn_idx == 0 else ""

        if c_lower in ("started_at", "created_at", "executed_at"):
            return row_data.get("started_at") if turn_idx == 0 else ""

        if c_lower in ("completed_at",):
            return row_data.get("completed_at") if turn_idx == 0 else ""

        if c_lower in ("correlation_id", "run_id", "id"):
            return row_data.get("correlation_id") if turn_idx == 0 else ""

        if c_lower in ("quality_score", "quality"):
            return row_data.get("quality_score") if turn_idx == 0 else ""

        if c_lower in ("safety_score", "safety"):
            return row_data.get("safety_score") if turn_idx == 0 else ""

        if c_lower in ("total_tokens", "tokens"):
            return row_data.get("total_tokens") if turn_idx == 0 else ""

        if c_lower in ("total_turns", "turns_count"):
            return row_data.get("total_turns") if turn_idx == 0 else ""

        if c_lower in ("error_message", "error"):
            return row_data.get("error_message") if turn_idx == 0 else ""

        # 8. Chat URL & Session Links (Shown on first row of scenario)
        if c_lower in ("chat_url", "chat_session_url", "session_url", "url", "chat_link", "chat_url_creator"):
            return (row_data.get("chat_url") or cvars.get("chat_url") or cvars.get("url") or dvars.get("chat_url") or "") if turn_idx == 0 else ""

        # Direct variable lookup in cvars or dvars
        if col_id in cvars:
            return cls._clean_response_payload(cvars[col_id]) if turn_idx == 0 else ""
        if col_id in dvars:
            return dvars[col_id] if turn_idx == 0 else ""

        return ""

    @classmethod
    def _clean_response_payload(cls, val: Any) -> Any:
        """Recursively unwraps JSON or dictionary into clean, human-readable text without raw JSON syntax."""
        if val is None:
            return ""
        if isinstance(val, dict):
            # Check standard assistant response fields in order of specificity
            if "answer" in val and val["answer"] is not None:
                return cls._clean_response_payload(val["answer"])
            if "response" in val and val["response"] is not None:
                return cls._clean_response_payload(val["response"])
            if "output" in val and val["output"] is not None:
                return cls._clean_response_payload(val["output"])
            if "result" in val and val["result"] is not None:
                return cls._clean_response_payload(val["result"])
            if "content" in val and val["content"] is not None:
                return cls._clean_response_payload(val["content"])
            if "text" in val and val["text"] is not None:
                return cls._clean_response_payload(val["text"])
            if "message" in val and isinstance(val["message"], str):
                return val["message"]
            return json.dumps(val, ensure_ascii=False, indent=2)
        if isinstance(val, str):
            trimmed = val.strip()
            if (trimmed.startswith("{") and trimmed.endswith("}")) or (trimmed.startswith("[") and trimmed.endswith("]")):
                try:
                    parsed = json.loads(trimmed)
                    if isinstance(parsed, (dict, list)):
                        return cls._clean_response_payload(parsed)
                except Exception:
                    pass

            # Regex unwrap fallback for {"answer": "...", ...} or {\n "answer": "..."}
            if '"answer"' in trimmed or "'answer'" in trimmed or '"response"' in trimmed:
                m = re.search(r'["\'](?:answer|response|content|text|output)["\']\s*:\s*["\'](.*?)["\']\s*(?:,|\}|$)', trimmed, re.DOTALL)
                if m:
                    raw_extracted = m.group(1)
                    clean = raw_extracted.replace('\\"', '"').replace('\\n', '\n').replace('\\t', '\t').strip()
                    if clean:
                        return clean

            return trimmed
        return val
