import pytest
import openpyxl
from app.domain.context import ExecutionContext
from app.execution.handlers.chat_url_handler import ChatUrlHandler
from app.execution.report_generator import ExcelReportGenerator
from app.domain.types import NodeType

@pytest.mark.asyncio
async def test_chat_url_handler_basic_interpolation():
    context = ExecutionContext(dataset_vars={"session_id": "sess-9941", "user_id": "usr-123"})
    
    config = {
        "base_url": "https://copilot.delphi.internal/chat",
        "query_template": "?id={session_id}&user={user_id}",
        "variable_name": "chat_url"
    }

    res = await ChatUrlHandler.execute(config, context)
    assert res["status"] == "SUCCESS"
    assert res["chat_url"] == "https://copilot.delphi.internal/chat?id=sess-9941&user=usr-123"
    assert context.get_variable("chat_url") == "https://copilot.delphi.internal/chat?id=sess-9941&user=usr-123"

@pytest.mark.asyncio
async def test_chat_url_handler_double_bracket_and_separators():
    context = ExecutionContext(dataset_vars={"session_id": "abc-7788"})
    
    # Query without leading ?
    config1 = {
        "base_url": "https://agent.ai/conversation",
        "query_template": "id={{session_id}}&env=production",
        "variable_name": "my_link"
    }
    res1 = await ChatUrlHandler.execute(config1, context)
    assert res1["chat_url"] == "https://agent.ai/conversation?id=abc-7788&env=production"
    assert context.get_variable("my_link") == "https://agent.ai/conversation?id=abc-7788&env=production"

    # Base URL with existing query param
    config2 = {
        "base_url": "https://agent.ai/conversation?app=travel",
        "query_template": "?id={session_id}",
        "variable_name": "chat_url"
    }
    res2 = await ChatUrlHandler.execute(config2, context)
    assert res2["chat_url"] == "https://agent.ai/conversation?app=travel&id=abc-7788"

@pytest.mark.asyncio
async def test_chat_url_in_excel_report():
    # Mock execution with CHAT_URL_CREATOR step
    class MockStep:
        def __init__(self, node_key, node_type, output_data):
            self.node_key = node_key
            self.node_type = node_type
            self.output_data = output_data

    class MockExecution:
        def __init__(self):
            self.id = "exec-chat-1"
            self.status = "PASSED"
            self.total_duration_ms = 450.0
            self.started_at = "2026-08-31T12:00:00Z"
            self.runtime_context = {
                "scenario_index": 1,
                "dataset_vars": {"message": "Hello Agent", "session_id": "sess-4455"}
            }
            self.steps = [
                MockStep("node-1", "PROMPT", {"message": "Hello Agent"}),
                MockStep("node-url", "CHAT_URL_CREATOR", {
                    "chat_url": "https://copilot.delphi.internal/chat?id=sess-4455",
                    "captured_variables": {"chat_url": "https://copilot.delphi.internal/chat?id=sess-4455"}
                })
            ]
            self.trace_events = []

    exec_item = MockExecution()
    
    template = {
        "columns": [
            {"id": "scenario_index", "label": "Scenario #", "enabled": True},
            {"id": "input_message", "label": "Initial Message", "enabled": True},
            {"id": "chat_url", "label": "Chat Session URL", "enabled": True},
            {"id": "status", "label": "Status", "enabled": True}
        ],
        "include_summary": True,
        "merge_scenario_cells": True
    }

    report_bytes = ExcelReportGenerator.generate_report("Test Chat Project", [exec_item], template)
    wb = openpyxl.load_workbook(report_bytes)
    assert "Detailed Scenarios" in wb.sheetnames
    
    ws = wb["Detailed Scenarios"]
    # Row 1 is header
    assert ws.cell(row=1, column=3).value == "Chat Session URL"
    # Row 2 is data with clickable hyperlink
    cell = ws.cell(row=2, column=3)
    assert cell.value == "https://copilot.delphi.internal/chat?id=sess-4455"
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://copilot.delphi.internal/chat?id=sess-4455"
