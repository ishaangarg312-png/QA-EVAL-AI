import io
import openpyxl
import pytest
from app.execution.report_generator import ExcelReportGenerator

class MockExecutionRun:
    def __init__(self, id, correlation_id, status, duration_ms, runtime_context, steps=None, trace_events=None):
        self.id = id
        self.correlation_id = correlation_id
        self.status = status
        self.total_duration_ms = duration_ms
        self.runtime_context = runtime_context
        self.steps = steps or []
        self.trace_events = trace_events or []
        self.started_at = "2026-08-31T12:00:00Z"
        self.created_at = "2026-08-31T12:00:00Z"
        self.completed_at = "2026-08-31T12:00:05Z"
        self.quality_score = 98.5
        self.safety_score = 100.0
        self.total_tokens = 1520
        self.error_message = None

def test_excel_report_generation_default():
    exec1 = MockExecutionRun(
        id="run-1",
        correlation_id="corr-matrix-job1-s1-abc",
        status="PASSED",
        duration_ms=2450.0,
        runtime_context={
            "scenario": "Explain work policies",
            "scenario_index": 1,
            "dataset_vars": {
                "message": "What is the hybrid policy?",
                "followup": "Can I work from home?",
                "message_api_response": "Delphi offers a hybrid work model with 2 days remote."
            },
            "turns": [
                {"message": "What is the hybrid policy?", "followup": "Can I work from home?"}
            ]
        }
    )

    exec2 = MockExecutionRun(
        id="run-2",
        correlation_id="corr-matrix-job1-s2-def",
        status="FAILED",
        duration_ms=1200.0,
        runtime_context={
            "scenario": "Ask for invalid document",
            "scenario_index": 2,
            "dataset_vars": {
                "message": "Show internal secret docs",
                "followup": "",
                "message_api_response": "Access denied."
            },
            "turns": [
                {"message": "Show internal secret docs", "followup": ""}
            ]
        }
    )

    buffer = ExcelReportGenerator.generate_report(
        project_name="Delphi Enterprise QA",
        executions=[exec1, exec2]
    )

    assert isinstance(buffer, io.BytesIO)
    wb = openpyxl.load_workbook(buffer)
    
    assert "Executive Summary" in wb.sheetnames
    assert "Detailed Scenarios" in wb.sheetnames

    ws_data = wb["Detailed Scenarios"]
    assert ws_data.max_row == 3 # Header + 2 data rows
    assert ws_data.max_column >= 5

    # Check scenario 1 row
    assert ws_data.cell(row=2, column=1).value == 1
    assert "polic" in str(ws_data.cell(row=2, column=2).value).lower()

    # Check summary sheet
    ws_summary = wb["Executive Summary"]
    assert "Delphi Enterprise QA" in str(ws_summary.cell(row=1, column=1).value)

def test_excel_report_custom_template():
    exec1 = MockExecutionRun(
        id="run-1",
        correlation_id="corr-matrix-job1-s1-abc",
        status="PASSED",
        duration_ms=3500.0,
        runtime_context={
            "scenario": "Multi-turn QA",
            "scenario_index": 1,
            "dataset_vars": {
                "message": "Hello",
                "custom_header": "Enterprise Value"
            },
            "turns": [
                {"message": "Hello", "followup": "How are you?"}
            ]
        }
    )

    custom_template = {
        "include_summary": True,
        "columns": [
            {"id": "scenario_index", "label": "Test ID", "enabled": True},
            {"id": "title", "label": "Scenario Description", "enabled": True},
            {"id": "status", "label": "Result Status", "enabled": True},
            {"id": "custom_header", "label": "Enterprise Header", "enabled": True},
            {"id": "duration_ms", "label": "Latency (ms)", "enabled": False} # Disabled
        ]
    }

    buffer = ExcelReportGenerator.generate_report(
        project_name="Custom Project",
        executions=[exec1],
        template_config=custom_template
    )

    wb = openpyxl.load_workbook(buffer)
    ws_data = wb["Detailed Scenarios"]
    
    assert ws_data.cell(row=1, column=4).value == "Enterprise Header"
    assert ws_data.cell(row=2, column=4).value == "Enterprise Value"

def test_excel_report_multi_turn_row_wise_expansion():
    class MockTrace:
        def __init__(self, title, payload):
            self.title = title
            self.raw_payload = payload

    # Exec 1: 3 Follow-up Questions
    exec_multi = MockExecutionRun(
        id="run-multi",
        correlation_id="corr-matrix-job1-s1-multi",
        status="PASSED",
        duration_ms=4500.0,
        runtime_context={
            "scenario": "Multi-turn interview",
            "scenario_index": 1,
            "dataset_vars": {
                "message": "Explain company policies",
                "followup": "What about remote work?",
                "message_api_response": "Delphi maintains policies for all teams."
            },
            "turns": [
                {"message": "Explain company policies", "followup": "What about remote work?"},
                {"followup": "What about vacation days?"},
                {"followup": "How to file expense reports?"}
            ]
        },
        trace_events=[
            MockTrace("Follow-up Questions (Turn #1)", {"answer": "Remote work is permitted 2 days a week."}),
            MockTrace("Follow-up Questions (Turn #2)", {"answer": "Vacation policy is 25 days annually."}),
            MockTrace("Follow-up Questions (Turn #3)", {"answer": "Expense reports are submitted through SAP."}),
        ]
    )

    # Exec 2: 0 Follow-up Questions (no follow-up node added)
    exec_zero_followup = MockExecutionRun(
        id="run-single",
        correlation_id="corr-matrix-job1-s2-single",
        status="PASSED",
        duration_ms=1200.0,
        runtime_context={
            "scenario": "Single Turn Query",
            "scenario_index": 2,
            "dataset_vars": {
                "message": "Hello There",
                "message_api_response": "Hello! How can I help you today?"
            },
            "turns": [
                {"message": "Hello There"}
            ]
        }
    )

    template = {
        "columns": [
            {"id": "scenario_index", "label": "Scenario #", "enabled": True},
            {"id": "input_message", "label": "Initial Message", "enabled": True},
            {"id": "response_message", "label": "Initial Response", "enabled": True},
            {"id": "input_followup", "label": "Follow-up Question", "enabled": True},
            {"id": "response_followup", "label": "Follow-up Response", "enabled": True},
            {"id": "status", "label": "Status", "enabled": True},
        ]
    }

    buffer = ExcelReportGenerator.generate_report(
        project_name="Multi-Turn Row-Wise Project",
        executions=[exec_multi, exec_zero_followup],
        template_config=template
    )

    wb = openpyxl.load_workbook(buffer)
    ws = wb["Detailed Scenarios"]

    # Scenario 1 - Turn 1 (Row 2 Top Merged Cell)
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Explain company policies"
    assert ws.cell(row=2, column=3).value == "Delphi maintains policies for all teams."
    assert ws.cell(row=2, column=4).value == "What about remote work?"
    assert ws.cell(row=2, column=5).value == "Remote work is permitted 2 days a week."
    assert ws.cell(row=2, column=6).value == "PASSED"

    # Scenario 1 - Turn 2 (Row 3 - Distinct Follow-up Question/Response)
    assert ws.cell(row=3, column=4).value == "What about vacation days?"
    assert ws.cell(row=3, column=5).value == "Vacation policy is 25 days annually."

    # Scenario 1 - Turn 3 (Row 4 - Distinct Follow-up Question/Response)
    assert ws.cell(row=4, column=4).value == "How to file expense reports?"
    assert ws.cell(row=4, column=5).value == "Expense reports are submitted through SAP."

    # Scenario 2 - Single turn (Row 5 - 0 follow-up questions)
    assert ws.cell(row=5, column=1).value == 2
    assert ws.cell(row=5, column=2).value == "Hello There"
    assert ws.cell(row=5, column=3).value == "Hello! How can I help you today?"
    assert not ws.cell(row=5, column=4).value
    assert not ws.cell(row=5, column=5).value
    assert ws.cell(row=5, column=6).value == "PASSED"

    # Verify Dynamic Cell Merges for Multi-turn Scenario 1
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    assert "A2:A4" in merged_ranges # Scenario # merged vertically
    assert "B2:B4" in merged_ranges # Initial Message merged vertically
    assert "C2:C4" in merged_ranges # Initial Response merged vertically
    assert "F2:F4" in merged_ranges # Status merged vertically
    # Follow-up Question and Response should NOT be merged
    assert "D2:D4" not in merged_ranges
    assert "E2:E4" not in merged_ranges
